"""MIT License

Copyright (c) 2020 Mario Mori

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE."""
"""
Procesa un reporte excel del modulo de reportes del sistema Scanntech
 y produce la exportación de ATER
"""

from openpyxl import load_workbook
from datetime import datetime, timedelta
from time import strftime
from afip.cuit import Cuit
from sys import exit
from sqlalchemy import create_engine
from sqlalchemy import Column, Integer, String
from sqlalchemy.orm import declarative_base, Session

from sqlalchemy.sql.expression import null

Base = declarative_base()
from typing import Generator

class OldAter(Base):
    """Old Model of ater"""
    __tablename__ = "old_ater"
    id = Column(Integer,primary_key=True)
    tipo_de_agente = Column(String(length=1),nullable=False)
    motivo_movimiento = Column(String(length=3),nullable=False)
    cuit_cliente = Column(String(length=11),nullable=False)
    fecha_percepcion = Column(String(length=10),nullable=False)
    tipo_de_comprobante = Column(String(length=6),nullable=False)
    letra_comprobante = Column(String(length=1),nullable=False)
    numero_de_comprobante = Column(String(length=12),nullable=False)
    importe_base = Column(String(length=15),nullable=False)
    alicuota = Column(String(length=6),nullable=False)
    importe_percibido = Column(String(length=15),nullable=False)
    anulacion = Column(String(length=1),nullable=False)
    contribuyente_conv_multi = Column(String(length=1),nullable=False)

    def __repr__(self):
        message = f"""
            id = {self.id},
            tipo_de_agente = {self.tipo_de_agente},
            motivo_movimiento = {self.motivo_movimiento},
            cuit_cliente = {self.cuit_cliente},
            fecha_percepcion = {self.fecha_percepcion},
            tipo_de_comprobante = {self.tipo_de_comprobante},
            letra_comprobante = {self.letra_comprobante},
            numero_comprobante = {self.numero_comprobante},
            importe_base = {self.importe_base},
            alicuota = {self.alicuota},
            importe_percibido = {self.importe_percibido},
            anulacion = {self.anulacion},
            contribuyente_conv_multi = {self.contribuyente_conv_multi}"""
        return message

    def line(self):
        return f"""{self.tipo_de_agente}{self.motivo_movimiento}{self.cuit_cliente}{self.fecha_percepcion}{self.tipo_de_comprobante}{self.letra_comprobante}{self.numero_de_comprobante}{self.importe_base}{self.alicuota}{self.importe_percibido}{self.anulacion}{self.contribuyente_conv_multi}\n"""

class NewAter(Base):
    """New Model of ater"""
    __tablename__ = "new_ater"
    id = Column(Integer,primary_key=True)
    tipo_de_comprobante = Column(String(length=3),nullable=False)
    letra_comprobante = Column(String(length=1),nullable=False)
    numero_de_comprobante = Column(String(length=12),nullable=False)
    cuit_cliente = Column(String(length=11),nullable=False)
    fecha_percepcion = Column(String(length=10),nullable=False)
    monto_sujeto_a_percepcion = Column(String(length=12),nullable=False)
    alicuota = Column(String(length=6),nullable=False)
    monto_percibido = Column(String(length=12),nullable=False)
    tipo_de_regimen_de_percepcion = Column(String(length=3),nullable=False)
    jurisdiccion = Column(String(length=3),nullable=False)

    def __repr__(self):
        message = f"""
            id = {self.id},
            tipo_de_comprobante = {self.tipo_de_comprobante},
            letra_comprobante = {self.letra_comprobante},
            numero_de_comprobante = {self.numero_de_comprobante},
            cuit_cliente = {self.cuit_cliente},
            fecha_percepcion = {self.fecha_percepcion},
            monto_sujeto_a_percepcion = {self.monto_sujeto_a_percepcion},
            alicuota = {self.alicuota},
            monto_percibido = {self.monto_percibido},
            tipo_de_regimen_de_percepcion = {self.tipo_de_regimen_de_percepcion},
            jurisdiccion = {self.jurisdiccion}"""
        return message

    def line(self):
        numero_renglon = str(self.id).zfill(5)
        return f"""{numero_renglon},{self.tipo_de_comprobante},{self.letra_comprobante},{self.numero_de_comprobante},{self.cuit_cliente},{self.fecha_percepcion},{self.monto_sujeto_a_percepcion},{self.alicuota},{self.monto_percibido},{self.tipo_de_regimen_de_percepcion},{self.jurisdiccion}\n"""


class Ater(object):
    """Manage the database and write of reports"""
    def __init__(self, *args):
        self.argv = args[0]
        if len(self.argv) > 2:
            if self.argv[1] == "old":
                self.old = True
        else:
            self.old = False
        self.excel_date = datetime(year=1900,month=1, day=1)
        self.cuit = Cuit()
        self.tipo_de_agente = "1"
        self.motivo = "061"
        self.alicuota = "003.00"
        self.anulacion = "0"
        self.contribuyente_conv_multi = "0"
        self.tipo_de_regimen_de_percepcion = "004"
        self.jurisdiccion = "908"
        engine = create_engine("sqlite+pysqlite:///:memory:")
        Base.metadata.create_all(engine)
        self.session = Session(engine)

    def help_app(self):
        msg = """\
                Ater necesita un archivo excel de donde puede sacar los datos en el modulo de reportes
                se encuentra el reporte \"ater\"
                """
        print(msg)
    def _parser_line(self,line):
        return [line[45:60],line[65:81],line[25:27]]

    def results_of_report(self, report_file):
        try:
            totals = {
                'monto_base' : 0.0,
                'monto_percibido' : 0.0
            }
            with open(report_file,mode="r",encoding="latin-1") as f:
                for line in f.readlines():
                    values_of_line = self._parser_line(line)
                    if values_of_line[2] == "TF":
                        totals['monto_base'] += float(values_of_line[0])
                        totals['monto_percibido'] += float((values_of_line[1]
                                                            .replace("\n","")
                                                            .replace("\'","")))
                    else:
                        totals['monto_base'] -= float(values_of_line[0])
                        totals['monto_percibido'] -= float((values_of_line[1]
                                                            .replace("\n","")
                                                            .replace("\'","")))
            print(totals)
        except Exception as e:
            print(e)
            raise e
        return

    def run(self):
        # Carga de excel
        try:
            if len(self.argv) > 2:
                wb = load_workbook(self.argv[2])
            else:
                wb = load_workbook(self.argv[1])
        except IndexError:
            self.help_app()
            exit()

        sheet = wb['0 - Tickets de Clientes con Fac']
        rows = sheet.rows
        next(rows)
        next(rows)
        next(rows)
        for row in rows:
            cuit_cliente = ""
            fecha_percepcion = ""
            tipo_de_comprobante = ""
            tipo_de_comprobante_new = ""
            letra_comprobante = ""
            numero_de_comprobante = ""
            importe_base = ""
            importe_percibido = ""
            # Verificamos si la celda esta en blanco se cierra el for
            if row[0].value == None:
                break
            # Revisamos que el cuit este bien si no se continua con la siguiente linea
            try:
                if self.cuit.verificador(str(row[8].value).replace(".","")):
                    cuit_cliente = str(row[8].value).replace(".","")
                else:
                    continue
            except TypeError:
                print("Ocurrio un error con la cuit {} del nro de operacion {}.\nError ocurrido: Tipo de formato".format(row[8].value,row[4].value))
            # Como obtener el valor de la fecha
            if row[0].value != "Fecha":
                try:
                    days = timedelta(days = (int(row[0].value) - 2))
                    fecha_percepcion = (self.excel_date + days).strftime("%d/%m/%Y")
                except TypeError:
                    dia = row[0].value
                    fecha_percepcion = dia.strftime("%d/%m/%Y")
            if row[1].value == "FACTURA":
                tipo_de_comprobante = "TF    "
                tipo_de_comprobante_new = "1"
            elif row[1].value == "NOTA DE CREDITO":
                tipo_de_comprobante = "C     "
                tipo_de_comprobante_new = "102"
            # Obtenemos la letra de la factura
            letra_comprobante = row[2].value
            # Obtenemos el punto de venta
            punto_venta = int(row[3].value)
            numero_de_comprobante += str(punto_venta).zfill(4)
            # Numero de comprobante
            numero_comprobante = row[4].value
            numero_de_comprobante += str(numero_comprobante).replace(".","").zfill(8)
            # Obtenemos el importe base
            total = row[5].value
            iva = row[6].value
            if total == 0.0:
                continue
            else:
                monto_base = float(total) - float(iva)
                importe_base = "{:.2f}".format(monto_base).zfill(15)
            # Obtenemos el importe percibido
            importe_percibido = row[7].value
            if importe_percibido == 0.0:
                continue
            else:
                importe_percibido = "{:.2f}".format(importe_percibido).zfill(15)
            # Agregamos el dato del convenio multilateral
            # Grabamos la linea en un archivo txt
            add_old_ater = OldAter(
                                tipo_de_agente = self.tipo_de_agente,
                                motivo_movimiento = self.motivo,
                                cuit_cliente = cuit_cliente,
                                fecha_percepcion = fecha_percepcion,
                                tipo_de_comprobante = tipo_de_comprobante,
                                letra_comprobante = letra_comprobante,
                                numero_de_comprobante = numero_de_comprobante,
                                importe_base = importe_base,
                                alicuota = self.alicuota,
                                importe_percibido = importe_percibido,
                                anulacion = self.anulacion,
                                contribuyente_conv_multi = self.contribuyente_conv_multi)
            add_new_ater = NewAter(
                                tipo_de_comprobante = tipo_de_comprobante_new,
                                letra_comprobante = letra_comprobante,
                                numero_de_comprobante = numero_de_comprobante,
                                cuit_cliente = cuit_cliente,
                                fecha_percepcion = fecha_percepcion,
                                monto_sujeto_a_percepcion = importe_base,
                                alicuota = self.alicuota,
                                monto_percibido = importe_percibido,
                                tipo_de_regimen_de_percepcion = self.tipo_de_regimen_de_percepcion,
                                jurisdiccion = self.jurisdiccion)
            self.session.add(add_old_ater)
            self.session.add(add_new_ater)
            self.session.flush()
        self.session.commit()
        print("Excel procesado")
        if self.old == True:
            with open("ater.txt",mode="w",newline="\r\n") as f:
                data = self.session.query(OldAter).order_by(OldAter.id).all()
                for d in data:
                    line = d.line()
                    f.write(line.replace("\'",""))
        else:
            with open("ater.txt",mode="w",newline="\r\n") as f:
                data = self.session.query(NewAter).order_by(NewAter.id).all()
                for d in data:
                    line = d.line()
                    f.write(line.replace("\'",""))
        print("Archivo ater.txt grabado")
        #self.results_of_report("ater.txt")

